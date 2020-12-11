# # Weekly Momentun Scanner Executed Weekly

import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from stock_config import nifty50_stock_list, nifty500_stock_list, momentum_stock_list, nse_fno_stock_list, sample_stock_list

def run_script(eval_date, stock_list, output_file):
    
    # ## Input the evaluation date below (Every Friday date), stock universe and output filename
    eval_date = eval_date
    stock_list = stock_list
    print(f"Total stocks in the list {len(stock_list)}")

    filename = output_file 

    # #### Function to check valid trading day. If open price is not returned then PREVIOUS date is checked until valid date is found.
    # ##### End date is valid date + 1. Used to get LTP for desired date using yfinance.

    
    def get_valid_date(start_date, end_date):
        valid = False
        while not valid:
            valid_check_data = yf.download('RELIANCE.NS', start = start_date, end=end_date)
            try:
                open_price = valid_check_data.iloc[0]['Open']
                valid = True
            except:
                start_date = (pd.to_datetime(start_date) - pd.DateOffset(days=1)).strftime('%Y-%m-%d')
                end_date = (pd.to_datetime(start_date) + pd.DateOffset(days=1)).strftime('%Y-%m-%d')
                valid = False
        return start_date, end_date

    
    # #### Get valid start and end dates for this week
    start_date=eval_date
    end_date = (pd.to_datetime(start_date) + pd.DateOffset(days=1)).strftime('%Y-%m-%d')
    start_date, end_date = get_valid_date(start_date, end_date)

    # ### Fetch this week data
    # #### [Below data fetch will throw error if there is no trading day or if the scrip doesnt have ltp (not listed) between start and end dates]
    data = yf.download(stock_list, start = start_date, end=end_date)
    this_week_close = pd.DataFrame(data.iloc[0]['Adj Close'])

    # #### Calculate dates for last week
    start_date=(pd.to_datetime(eval_date) - pd.DateOffset(days=7)).strftime('%Y-%m-%d')
    end_date = (pd.to_datetime(start_date) + pd.DateOffset(days=1)).strftime('%Y-%m-%d')

    # #### Get valid start and end dates for last week
    start_date, end_date = get_valid_date(start_date, end_date)

    # ### Fetch last week data
    # #### [Below data fetch will throw error if there is no trading day between start and end dates]
    data = yf.download(stock_list, start = start_date, end=end_date)
    last_week_close = pd.DataFrame(data.iloc[0]['Adj Close'])
    
    final_df = last_week_close.join(this_week_close)
    
    # #### Calculate dates for Next Week (For Forward Testing)
    start_date=(pd.to_datetime(eval_date) + pd.DateOffset(days=7)).strftime('%Y-%m-%d')
    end_date = (pd.to_datetime(start_date) + pd.DateOffset(days=1)).strftime('%Y-%m-%d')

    # #### Get valid start and end dates for next week
    start_date, end_date = get_valid_date(start_date, end_date)
    
    # ### Fetch next week data
    # #### [Below data fetch will throw error if there is no trading day between start and end dates]
    data = yf.download(stock_list, start = start_date, end=end_date)
    next_week_close = pd.DataFrame(data.iloc[0]['Adj Close'])
    final_df = final_df.join(next_week_close)

    final_df.rename(columns =lambda t: t.strftime('%Y-%m-%d'), inplace=True)

    
    # ## Calculate Momentum
    last_week_date = final_df.columns[0]
    this_week_date = final_df.columns[1]
    next_week_date = final_df.columns[2]
    final_df['Change(C-B)'] = final_df[this_week_date] - final_df[last_week_date]
    final_df['% Change'] = (final_df['Change(C-B)'] / final_df[last_week_date])*100
    final_df['Gain(D-B)'] = final_df[next_week_date] - final_df[this_week_date]
    final_df['% Gain'] = (final_df['Gain(D-B)'] / final_df[this_week_date])*100

    # ## Final Output
    final_df.sort_values(by=['% Change'], ascending=False, inplace=True)
    final_df.rename_axis('MOMENTUM_ATH', inplace=True)
    final_df.rename_axis('', axis='columns', inplace=True)

    # ### Create an empty excel file as below filename before running below code
    book = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine = 'openpyxl')
    writer.book = book

    final_df.round(2).to_excel(writer, sheet_name=eval_date)
    writer.save()
    writer.close()

    sheet = book.get_sheet_by_name(eval_date) 
    sheet['J20'] = '=SUM(C2:C20)'  
    sheet['K20'] = '=SUM(G2:G20)'
    sheet['M20'] = '=K20/J20'
    sheet['M20'].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE_00

    for column_cells in sheet.columns:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = 12    
    sheet.column_dimensions['B'].width = 20

    book.save(filename)  
    print(f"Done. Output file path and name --> {filename}")


if __name__ == "__main__":
    eval_date = "2020-01-31"
    script_end_date = "2020-02-06"
    stock_list = momentum_stock_list
    output_file = r'data/output/momentum_weekly_output.xlsx'

    while (eval_date <= script_end_date):
        run_script(eval_date,stock_list,output_file)
        eval_date = (pd.to_datetime(eval_date) + pd.DateOffset(days=7)).strftime('%Y-%m-%d')
